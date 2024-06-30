import os
import sys
import socket
from utils import getpaths



import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
from utils import d2s, stripper, hasinput
import datetime
from datetime import timedelta


#Handle the input arguments from script file
try:
    scac = sys.argv[1]
    nt = 'remote'
except:
    print('Must have at least one argument...FELA or OSLM or NEVO')
    scac = 'oslm'
    nt = 'remote'

#Handle the input arguments from script file
try:
    daysback = sys.argv[2]
    try:
        daysback = int(daysback)
    except:
        print('Error the second argument must be an interger')
        quit()
except:
    print('Using default of 365 days back')
    daysback = 2023

if daysback > 2000 and daysback < 2025:
    yearby = 1
    d1 = datetime.date(daysback, 1, 1)
    d2 = datetime.date(daysback, 12, 31)
else:
    yearby = 0

if yearby:
    if daysback == 2022:
        payroll_start = datetime.date(daysback-1, 12, 26)
        payroll_stop = datetime.date(daysback, 12, 31)
    if daysback == 2023:
        payroll_start = datetime.date(daysback-1, 12, 24)
        payroll_stop = datetime.date(daysback, 12, 31)
    if daysback == 2024:
        payroll_start = datetime.date(daysback-1, 12, 31)
        payroll_stop = datetime.date(daysback, 12, 31)
    thisdate = payroll_start
    payroll = []
    print(payroll)
    for ix in range(28):
        if thisdate > payroll_stop: break
        thisbiweek = []
        for jx in range(14):
            thisdate = thisdate + timedelta(1)
            thisbiweek.append(thisdate)
        print(thisbiweek)
        payroll.append(thisbiweek)
    print(payroll)


today = datetime.datetime.today()
today_str = today.strftime("%m/%d/%Y")
d = today.strftime("%B %d, %Y")
cutoff = datetime.datetime.now() - timedelta(daysback)
cutoff = cutoff.date()
over30 = datetime.datetime.now() - timedelta(30)
over30 = over30.date()
todaydate = today.date()

scac = scac.upper()

if scac == 'OSLM' or scac == 'FELA' or scac == 'NEVO':

    print(f'Running FFF_dray_ARhistory for {scac}')
    host_name = socket.gethostname()
    print("Host Name:", host_name)
    dropbox_path = getpaths(host_name, 'dropbox')
    if yearby:
        ar_path = f'{dropbox_path}/Dray/{scac}_AR_History_{daysback}.xlsx'
    else:
        ar_path = f'{dropbox_path}/Dray/{scac}_AR_History_{todaydate}_{daysback}.xlsx'

    sys_path = getpaths(host_name, 'system')
    sys.path.append(sys_path) #So we can import CCC_system_setup from full path

    os.environ['SCAC'] = scac
    os.environ['PURPOSE'] = 'script'
    os.environ['MACHINE'] = host_name
    os.environ['TUNNEL'] = nt

    from remote_db_connect import db
    if nt == 'remote': from remote_db_connect import tunnel
    from models8 import Orders, Interchange, Invoices, Openi

else:
    scac = 'nogo'
    print('The argument must be FELA or OSLM or NEVO')
    quit()

def get_sector(dti, payroll):
    for ix, sector in enumerate(payroll):
        d1 = sector[0]
        d2 = sector[13]
        if dti >= d1 and dti <= d2:
            print(f'For dti {dti} returning sector {ix} with d1 {d1} and d2 {d2}')
            return ix
    return 9999

if scac != 'nogo':
    try:
        wb = openpyxl.load_workbook(ar_path, data_only=True)
    except:
        wb = openpyxl.Workbook()

    shtlist = wb.sheetnames
    for sht in shtlist:
        if sht != 'AR Notes':
            std = wb.get_sheet_by_name(sht)
            wb.remove_sheet(std)

    sheet_str = today_str.replace('/',' ')
    newsheet = f'AR History {sheet_str}'
    dfs= wb.create_sheet(newsheet)
    newsheet = f'AR by Date {sheet_str}'
    dfc= wb.create_sheet(newsheet)

    #formats for writing to excel
    money = '$#,##0.00'
    dec2 = '#,##0.00'
    dec0 = '#,##0'

    #Create column width calculation function
    def column_wide(headers,ydata):
        column_widths = []
        for cell in enumerate(headers):
            cell = str(cell)
            column_widths.append(len(cell))
        for row in ydata:
            for i, cell in enumerate(row):
                test = len(str(cell))
                column_widths[i] = max(test,column_widths[i])
        return column_widths

    #Write the header for AR by Date Sheet
    row = 1
    ydata = []
    hdrs = ['SID','JO','Company', 'Order', 'Release','Booking In','Container','Date Out','Date In','Date Invoiced', 'Amount All', 'Amount Chassis', 'Amount Non-Chassis', 'Amount Paid', 'Invoice Filename', 'Comment ID']
    for col, hdr in enumerate(hdrs):
        d = dfc.cell(row=row, column=col + 1, value=hdr)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
    success = 0
    trys = 0
    while success == 0 and trys < 20:
        try:
            if yearby:
                odata = Orders.query.filter((Orders.Istat > 0) & (Orders.Date >= d1) & (Orders.Date <= d2)).all()
            else:
                odata = Orders.query.filter((Orders.Istat>0)  & (Orders.Date > cutoff)).all()
            success = 1
        except:
            print(f'Could not open tunnel on try {trys}')
            trys = trys + 1

    if success == 1:
        #Make a list of each unique company with open invoices
        comps = []
        for odat in odata:
            comp = odat.Shipper
            if comp not in comps: comps.append(comp)

        for gdat in odata:
            print(gdat.id, gdat.Jo, gdat.Date)
            amtall = 0.00
            amtchas = 0.00
            amtother = 0.00
            amtpaid = 0.00
            jo = gdat.Jo
            istat = gdat.Istat
            con = gdat.Container
            if istat == 5 or istat == 8:
                paid = 1
            else:
                paid = 0
            links = gdat.Links
            odr = gdat.Order
            if links is None: links = ''
            invo = gdat.Invoice
            if invo is None: invo = ''

            obk = gdat.Booking
            obol = gdat.BOL
            if hasinput(obol):
                ietyp = 'imp'
            else:
                ietyp = 'exp'


            int_data = Interchange.query.filter(Interchange.Jo == jo).all()
            if len(int_data) >= 1:
                bk1 = int_data[0].Release
                con = int_data[0].Container
                dt1 = int_data[0].Date
            else:
                bk1 = 'None'
                dt1 = None
            if len(int_data) >= 2:
                bk2 = int_data[1].Release
                dt2 = int_data[1].Date
            else:
                bk2 = 'None'
                dt2 = None

            if ietyp == 'imp': bk1 = obol

            if bk1 == 'DP OP' or bk2 == 'DP OP':
                print(f'This booking {bk1} and {bk2} has DP OP')
            else:
                idat = Invoices.query.filter(Invoices.Jo == jo).first()
                if idat is not None:
                    dti = idat.Date
                    amtall = float(idat.Total)
                    kdat = Invoices.query.filter((Invoices.Jo == jo) & (Invoices.Service.contains('Chassis'))).first()
                    if kdat is not None:
                        amtchas = float(kdat.Amount)
                        amtother = amtall - amtchas
                    else:
                        amtchas = 0.00
                        amtother = amtall - amtchas
                    if paid:
                        amtpaid = amtall
                    else:
                        antpaid = 0.00
                    newblock = [gdat.id, gdat.Jo, gdat.Shipper, odr, bk1, bk2, con, dt1, dt2, dti, amtall, amtchas, amtother, amtpaid, invo, links]
                    ydata.append(newblock)
                else:
                    dti = 'None'
                    amt = 'No Invoice'
                    print(f'No invoice found for {gdat.id} with JO: {jo}')

        ydata.sort(key=lambda row: (row[9], row[12]))

        if yearby:
            newtab = f'Income Biweekly {sheet_str}'
            dpay = wb.create_sheet(title=newtab)

            #Create the amounts by payroll section
            m = len(payroll)
            pdata = [None] * m
            ptotal = [0.00] * m
            print(f'Len of payroll is {len(payroll)} and dafault pdata is {pdata}')
            for ydat in ydata:
                dti = ydat[9]
                this_sector = get_sector(dti, payroll)
                if this_sector > m:
                    print(f'Returned invalid sector {this_sector}')
                else:
                    if pdata[this_sector] == None: pdata[this_sector] = []
                    pdata[this_sector].append(ydat)
            for kx, pdat in enumerate(pdata):
                if pdat is not None:
                    print('')
                    print(f'This is for payroll number {kx} date {payroll[kx][0]} to {payroll[kx][13]} ')
                    for dat in pdat:
                        ptotal[kx] = ptotal[kx] + dat[10]
                        print(f'Company {dat[2]}  Date {dat[9]}  Amount {dat[10]}')

            print(ptotal)

            for kx, pdat in enumerate(pdata):
                if pdat is not None:
                    newtab = f'Payroll {kx}'
                    print(newtab)
                    ws = wb.create_sheet(title=newtab)
                    row = 1
                    title = f'For payroll {kx} from {payroll[kx][0]} to {payroll[kx][13]}'
                    d = ws.cell(row=row, column=1, value=title)
                    d.alignment = Alignment(horizontal='left')
                    d.font = Font(name='Calibri', size=10, bold=True)

                    row = 2



                    # write the headers
                    hdrs = ['SID', 'JO', 'Company', 'Order', 'Release', 'Booking In', 'Container', 'Date Out',
                            'Date In',
                            'Date Invoiced', 'Amount All', 'Amount Chassis', 'Amount Non-Chassis', 'Amount Paid',
                            'Invoice Filename', 'Comment ID']
                    for col, hdr in enumerate(hdrs):
                        d = ws.cell(row=row, column=col + 1, value=hdr)
                        d.alignment = Alignment(horizontal='center')
                        d.font = Font(name='Calibri', size=10, bold=True)

                    for dat in pdat:
                        shp = dat[2]
                        row = row + 1
                        for col, item in enumerate(dat):
                            d = ws.cell(row=row, column=col + 1, value=item)
                            d.alignment = Alignment(horizontal='center')
                            d.font = Font(name='Calibri', size=10, bold=False)
                            if col >= 10 and col <= 13: d.number_format = money

                    column_widths = column_wide(hdrs, ydata)
                    for i, column_width in enumerate(column_widths):
                        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

            #Write out the biweekly summary
            p_hdrs = ['Payroll #', 'From Date', 'To Date', 'Amounts', 'Commission']
            for col, hdr in enumerate(p_hdrs):
                d = dpay.cell(row=1, column=col + 2, value=hdr)
                d.alignment = Alignment(horizontal='center')
                d.font = Font(name='Calibri', size=10, bold=True)

            row_s = 1

            for kx, pdat in enumerate(pdata):
                if pdat is not None:
                    row_s += 1

                    d = dpay.cell(row=row_s, column=2, value=kx)
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=False)

                    d = dpay.cell(row=row_s, column=3, value=payroll[kx][0])
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=False)

                    d = dpay.cell(row=row_s, column=4, value=payroll[kx][13])
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=False)

                    d = dpay.cell(row=row_s, column=5, value=ptotal[kx])
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=False)
                    d.number_format = money

                    d = dpay.cell(row=row_s, column=6, value=ptotal[kx]*.075)
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=False)
                    d.number_format = money

            column_widths = column_wide(p_hdrs, [])
            for i, column_width in enumerate(column_widths):
                dpay.column_dimensions[get_column_letter(i + 2)].width = column_width + 4



        for ydat in ydata:
            row = row + 1
            dti = ydat[9]

            for col, item in enumerate(ydat):
                d = dfc.cell(row=row, column=col + 1, value=item)
                d.alignment = Alignment(horizontal='center')
                if dti < over30:
                    d.font = Font(name='Calibri', size=10, bold=False)
                else:
                    d.font = Font(name='Calibri', size=10, bold=False)

                if col >= 10 and col <= 13: d.number_format = money

        column_widths = column_wide(hdrs, ydata)
        for i, column_width in enumerate(column_widths):
            dfc.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

        #dfc.auto_filter.ref = f'A1:M{row}'
        #dfc.auto_filter.add_filter_column(8, ['Inactive'])
        #dfc.auto_filter.add_sort_condition(f'I2:I{row}')
        sumall_all = 0.0
        sumchas_all = 0.0
        sumother_all = 0.0
        sumpaid_all = 0.0
        zdata = []
        s_hdrs = ['Company', 'Total Invoiced', 'Chassis', 'Non-Chassis', 'Paid']
        desc = f'Report last updated on {todaydate}'
        #Kill all then entries and add them each in
        Openi.query.delete()
        db.session.commit()

        for col, hdr in enumerate(s_hdrs):
            d = dfs.cell(row=1, column=col + 2, value=hdr)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=True)

        #Write tabs for each company
        #exclude_these = ['FEL Ocean Division', 'First Eagle Logistics', 'One Stop Logistics', 'Nello Enterprise LLC', 'Jays Auto Service']
        exclude_these = []
        for comp in comps:
            if not any(comp in x for x in exclude_these):
                tab = comp.split()
                if len(tab)>1: newtab = f'{tab[0]} {tab[1]}'
                else: newtab = f'{tab[0]}'
                newtab = newtab.replace(',','')
                print(newtab)

                # Set up variables to sum the information for each company
                sumall = 0.0
                sumchas = 0.0
                sumother = 0.0
                sumpaid = 0.0

                ws = wb.create_sheet(title=newtab)
                row = 1
                # write the headers
                hdrs = ['SID', 'JO', 'Company', 'Order', 'Release', 'Booking In', 'Container', 'Date Out', 'Date In',
                        'Date Invoiced', 'Amount All', 'Amount Chassis', 'Amount Non-Chassis', 'Amount Paid', 'Invoice Filename', 'Comment ID']
                for col, hdr in enumerate(hdrs):
                    d = ws.cell(row=row, column=col + 1, value=hdr)
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=True)

                for ydat in ydata:
                    shp = ydat[2]
                    if comp == shp:
                        row = row + 1
                        sumall += ydat[10]
                        sumchas += ydat[11]
                        sumother += ydat[12]
                        sumpaid += ydat[13]

                        for col, item in enumerate(ydat):
                            d = ws.cell(row=row, column=col + 1, value=item)
                            d.alignment = Alignment(horizontal='center')
                            d.font = Font(name='Calibri', size=10, bold=False)
                            if col >= 10 and col <= 13: d.number_format = money

                for i, column_width in enumerate(column_widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

                #Write the Summary Data for Company to Summary Page
                zdata.append([comp,sumall,sumchas,sumother,sumpaid])
                sumall_all += sumall
                sumchas_all += sumchas
                sumother_all += sumother
                sumpaid_all += sumpaid


        row_s = 1
        zdata.sort(key=lambda row: row[1], reverse=True)
        for zdat in zdata:
            comp, sall, schas, soth, spaid = zdat
            compu = comp.upper()

            row_s += 1

            d = dfs.cell(row=row_s, column=2, value=comp)
            d.alignment = Alignment(horizontal='left')
            d.font = Font(name='Calibri', size=10, bold=False)

            d = dfs.cell(row=row_s, column=3, value=sall)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

            d = dfs.cell(row=row_s, column=4, value=schas)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

            d = dfs.cell(row=row_s, column=5, value=soth)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

            d = dfs.cell(row=row_s, column=6, value=spaid)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

        row_s = row_s + 2
        d = dfs.cell(row=row_s, column=2, value='Totals:')
        d.alignment = Alignment(horizontal='right')
        d.font = Font(name='Calibri', size=10, bold=True)

        d = dfs.cell(row=row_s, column=3, value=sumall_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        d = dfs.cell(row=row_s, column=4, value=sumchas_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        d = dfs.cell(row=row_s, column=5, value=sumother_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        d = dfs.cell(row=row_s, column=6, value=sumpaid_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        column_widths = column_wide(s_hdrs, zdata)
        for i, column_width in enumerate(column_widths):
            dfs.column_dimensions[get_column_letter(i + 2)].width = column_width + 4

        wb.save(ar_path)

tunnel.stop()