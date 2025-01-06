import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
import datetime
from datetime import timedelta
import socket
from utils import getpaths

today = datetime.datetime.today()
today_str = today.strftime("%m/%d/%Y")
d = today.strftime("%B %d, %Y")
print(today)
cutoff = datetime.datetime.now() - timedelta(360)
cutoff = cutoff.date()
over30 = datetime.datetime.now() - timedelta(30)
over30 = over30.date()

#Handle the input arguments from script file
try:
    scac = sys.argv[1]
except:
    print('Must have at least one argument...FELA or OSLM or NEVO')
    scac = 'oslm'

scac = scac.upper()
nt = 'remote'

if scac == 'OSLM' or scac == 'FELA' or scac == 'NEVO':
    print(f'Running FFF_dray_Global for {scac}')
    host_name = socket.gethostname()
    print("Host Name:", host_name)
    dropbox_path = getpaths(host_name, 'dropbox')
    ar_path = f'{dropbox_path}/Dray/{scac}_Global_Invoice_Report.xlsx'
    sys_path = getpaths(host_name, 'system')
    sys.path.append(sys_path) #So we can import CCC_system_setup from full path

    os.environ['SCAC'] = scac
    os.environ['PURPOSE'] = 'script'
    os.environ['MACHINE'] = host_name
    os.environ['TUNNEL'] = nt

    from remote_db_connect import tunnel, db
    from models8 import Orders, Interchange, Invoices, SumInv

else:
    scac = 'nogo'
    print('The argument must be FELA or OSLM or NEVO')
    quit()

if scac != 'nogo':

    #Get all Global Open Invoices
    wb = openpyxl.load_workbook(ar_path, data_only=True)
    sheet_str = today_str.replace('/',' ')
    newsheet = f'{sheet_str}'
    print(newsheet)
    dfc= wb.create_sheet(newsheet)

    #Create column width calculation function
    def column_wide(headers,ydata):
        column_widths = []
        for cell in enumerate(headers):
            cell = str(cell)
            column_widths.append(len(cell))
        for row in ydata:
            for i, cell in enumerate(row):
                test = len(str(cell))
                column_widths[i] = max(test,column_widths[i])
        return column_widths

    #Write the header
    row = 1
    ydata = []
    hdrs = ['SID','JO','Company','Booking Out','Booking In','Container','Date Out','Date In','Date Invoiced', 'Amount','Label', 'Comment ID']
    for col, hdr in enumerate(hdrs):
        d = dfc.cell(row=row, column=col + 1, value=hdr)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
    success = 0
    trys = 0
    while success == 0 and trys < 20:
        try:
            gdata = Orders.query.filter((Orders.Shipper == 'Global Business Link') & ((Orders.Istat<4) | (Orders.Istat == 6) | (Orders.Istat == 7)) & (Orders.Date > cutoff)).all()
            success = 1
        except:
            print(f'Could not open tunnel on try {trys}')
            trys = trys + 1

    if success == 1:
        siopen = []
        other_open = []
        for gdat in gdata:
            bk1 = 'None'
            bk2 = 'None'
            con = 'None'
            dt1 = None
            dt2 = None
            print(gdat.id, gdat.Date)
            jo = gdat.Jo
            links = gdat.Links
            label = gdat.Label
            if label is None: label = gdat.Order
            if label is not None:
                if 'SI' in label:
                    if label not in siopen: siopen.append(label)
                else:
                    label = gdat.Order
                    other_open.append(jo)
            if links is None: links = ''
            int_data = Interchange.query.filter(Interchange.Jo == jo).all()
            if len(int_data) >= 1:
                bk1 = int_data[0].Release
                con = int_data[0].Container
                dt1 = int_data[0].Date
            if len(int_data) >= 2:
                bk2 = int_data[1].Release
                dt2 = int_data[1].Date

            if bk1 == 'DP OP' or bk2 == 'DP OP':
                print(f'This booking {bk1} and {bk2} has DP OP')
            else:
                idat = Invoices.query.filter(Invoices.Jo == jo).first()
                if idat is not None:
                    dti = idat.Date
                    amt = idat.Total
                    newblock = [gdat.id, gdat.Jo, gdat.Shipper, bk1, bk2, con, dt1, dt2, dti, amt, label, links]
                    ydata.append(newblock)
                else:
                    dti = 'None'
                    amt = 'No Invoice'
                    print(f'No invoice found for {gdat.id}')

        ydata.sort(key=lambda row: (row[8], row[11]))

        for ydat in ydata:
            row = row + 1
            dti = ydat[8]
            for col, item in enumerate(ydat):
                d = dfc.cell(row=row, column=col + 1, value=item)
                d.alignment = Alignment(horizontal='center')
                if dti < over30:
                    d.font = Font(name='Calibri', size=10, bold=True)
                else:
                    d.font = Font(name='Calibri', size=10, bold=False)

        column_widths = column_wide(hdrs, ydata)
        print(column_widths)
        for i, column_width in enumerate(column_widths):
            dfc.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

        #dfc.auto_filter.ref = f'A1:M{row}'
        #dfc.auto_filter.add_filter_column(8, ['Inactive'])
        #dfc.auto_filter.add_sort_condition(f'I2:I{row}')
        wb.save(ar_path)

    #Run the summary invoice assessment:
    print(siopen)
    for si in siopen:
        sdata = SumInv.query.filter(SumInv.Si == si).all()
        num_sum = len(sdata)
        num_open = 0
        jolist_open = []
        jolist_paid = []
        jolist_gone = []
        for sdat in sdata:
            jo = sdat.Jo
            idat = Orders.query.filter(Orders.Jo == jo).first()
            if idat is not None:
                istat = idat.Istat
                if istat != 8:
                    jolist_open.append(idat.Jo)
                    num_open += 1
                else:
                    jolist_paid.append(idat.Jo)
            else:
                jolist_gone.append(jo)
        print('')
        print(f'Summary Invoice {sdat.Si} totaling {sdat.Total} emailed on {sdat.Date}')
        print(f'There are {num_sum} jobs on this summary invoice and {num_open} are still open')
        if num_sum != num_open: print(f'*********** Need to Investigate.  Why are {num_sum} and {num_open} not same *********')
        print(f'The jobs still open are: {jolist_open}')
        print(f'The jobs paid are: {jolist_paid}')
        print(f'The jobs gone/missing are: {jolist_gone}')
        print('')

    for jo in other_open:
        #idat = Orders.query.filter(Orders.Jo == jo).first()
        idat = Invoices.query.filter(Invoices.Jo == jo).first()
        if idat is not None:
            odat = Orders.query.filter(Orders.Jo == jo).first()
            if odat is not None:
                print(f'Regular invoice open for JO: {jo} ({odat.Order}|{odat.Booking}|{odat.Container} Total {idat.Total} emailed on {idat.Date}')

    tunnel.stop()