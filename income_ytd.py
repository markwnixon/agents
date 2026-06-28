import os
import sys
import socket
from utils import getpaths, hasinput, d2s, stripper

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
import datetime
from datetime import timedelta
import numpy as np

try:
    scac = sys.argv[1]
    nt = 'remote'
except:
    print('Must have at least one argument...FELA or OSLM or NEVO')
    scac = 'OSLM'
    nt = 'remote'

scac = scac.upper()

if scac == 'OSLM' or scac == 'FELA' or scac == 'NEVO':

    print(f'Running Income Report for {scac}')
    host_name = socket.gethostname()
    print("Host Name:", host_name)
    dropbox_path = getpaths(host_name, 'dropbox')
    ar_path = f'{dropbox_path}/Dray/{scac}_AR_Report.xlsx'
    sys_path = getpaths(host_name, 'system')
    sys.path.append(sys_path) #So we can import CCC_system_setup from full path

    os.environ['SCAC'] = scac
    os.environ['PURPOSE'] = 'script'
    os.environ['MACHINE'] = host_name
    os.environ['TUNNEL'] = nt

    from remote_db_connect import db
    if nt == 'remote': from remote_db_connect import tunnel
    from models8 import Orders, Drops, Gledger, Invoices, Interchange, Income, People

else:
    scac = 'nogo'
    print('The argument must be FELA or OSLM or NEVO')
    quit()

# Date Section: Define Dates of Interest
# Set up calculation of work days, holidays, etc so that we can make projections
################################################################################
################ Define the Year to Date Stop Point ############################
year, ytd_month, ytd_day = 2024, 7, 27
ytd = datetime.date(year, ytd_month, ytd_day)
################################################################################
categories = ['Line Item Total', 'Line Haul', 'Chassis Fees', 'Detention', 'Storage', 'Demurrage', 'Loading Charge', 'Towing', 'Other']

def last_day_of_month(any_day):
    # this will never fail
    # get close to the end of the month for any day, and add 45224 days 'over'
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    # subtract the number of remaining 'overage' days to get last day of current month, or said programattically said, the previous day of the first of next month
    return next_month - datetime.timedelta(days=next_month.day)

#Determine if the ytd month is a full month or part month
if ytd == last_day_of_month(ytd): fullmonth = True
else: fullmonth = False

#Create the invoice periods of interest. This creates a data output for each 2 weeks
nextyear = year + 1
d1 = datetime.date(2024, 1, 1)
thisyear = d1.year
keydates = []
while thisyear < nextyear:
    thisyear = d1.year
    keydates.append(d1)
    d1 = d1 + timedelta(7)
print(keydates)

#Define the port holiday schedule
port_holidays = [datetime.date(2024, 1, 1),
                 datetime.date(2024, 1, 15),
                 datetime.date(2024, 2, 19),
                 datetime.date(2024, 3, 25),
                 datetime.date(2024, 5, 27),
                 datetime.date(2024, 6, 19),
                 datetime.date(2024, 7, 4),
                 datetime.date(2024, 9, 2),
                 datetime.date(2024, 10, 14),
                 datetime.date(2024, 11, 11),
                 datetime.date(2024, 11, 28),
                 datetime.date(2024, 12, 25)]
holidays_yr = len(port_holidays)
# Calculate the number of workdays to date and for the year
today = datetime.datetime.today()
firstday = datetime.date(year, 1, 1)
lastday = datetime.date(year, 12, 31)
ytd = datetime.date(year, ytd_month, ytd_day)
monlist = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
weekdays_yr = np.busday_count(firstday, lastday)
weekdays_ytd = np.busday_count(firstday, ytd)
holidays_past = [y for y in port_holidays if y <= ytd]
holidays_ytd = len(holidays_past)
holidays_future = holidays_yr - holidays_ytd
workdays_yr = weekdays_yr - holidays_yr
workdays_ytd = weekdays_ytd - holidays_ytd
print(f'For the YEAR there are {weekdays_yr} weekdays, {holidays_yr} holidays, and {workdays_yr} work days')
print(f'For the year to date as of {ytd} there are {weekdays_ytd} weekdays, {holidays_ytd} holidays, and {workdays_ytd} work days')
proj_fraction = workdays_yr/workdays_ytd

#Separate customers into invoiced and cash-paying, this provides a way to determine what
#amounts should show up in bank
legitco = []
cashco = []

success = 0
trys = 0
while success == 0 and trys < 20:
    try:
        pdata = People.query.filter(People.Ptype=='Trucking').all()
        success = 1
    except:
        print(f'No success opening tunnel on try {trys}')
        success = 0
        trys = trys + 1

if success:

    for pdat in pdata:
        e1 = pdat.Email
        e2 = pdat.Associate1
        e3 = pdat.Associate2
        if hasinput(e1) or hasinput(e2) or hasinput(e3):
            legitco.append(pdat.Company)
        else:
            cashco.append(pdat.Company)
    #print(f'Legitco: {legitco}')
    #print(f'Cashco : {cashco}')



    # Clean up interchange.  Sometimes interchange does not get correct company information and this cleans that
    nays = Interchange.query.filter(((Interchange.Company == 'NAY') | (Interchange.Company == 'Ocean')) & (Interchange.Status == 'IO') & (Interchange.Date >= firstday) ).all()
    for nay in nays:
        co = nay.Company
        con = nay.Container
        rel = nay.Release
        ord = Orders.query.filter((Orders.Container == con) & (Orders.Booking == rel)).first()
        if ord is not None:
            print(f'Fixing NAY and Ocean for {con} {rel} to {ord.Jo} {ord.Shipper}')
            nay.Company = ord.Shipper
        else:
            print(f'No Order for NAY and Ocean for {con} {rel} on {nay.Date}')
    db.session.commit()
    print('Completed cleaning up the Interchange database')

    # Calculate the port days in range of dates of interest, beginning of year to end of period of interest
    # Also create the unique items to find:
    # 1) Make sure every container has an associated job or else report that
    # 2) Make sure every job has a container or report that
    # 3)
    port_days = []
    conid_in = []
    conid_out_only = []
    conid_nojob = []
    print(f'Grabbing containers in range {firstday} to {ytd}')
    cdata_all = Interchange.query.filter( (Interchange.Date >= firstday) & (Interchange.Date <= ytd) ).order_by(Interchange.Date).all()
    for cdat in cdata_all:
        next_date = cdat.Date
        if next_date not in port_days:
            print(f'Need to add date {next_date} to port days')
            port_days.append(next_date)
    num_port_days = len(port_days)
    port_day_tags = [[] for _ in range(num_port_days)]
    port_day_numtrucks = [0] * num_port_days
    print(f'Completed port days calcs.  There are {num_port_days} port days from {firstday} to {ytd}')
    print(port_days)
    print(port_day_tags)
    # Now find how many trucks we ran for each port day
    for cdat in cdata_all:
        cdate = cdat.Date
        tag = cdat.TruckNumber
        if tag is None: tag = 'NoTag'
        if len(tag) == 6:
            iport = port_days.index(cdate)
            port_day_tags[iport].append(tag)
    for jx, tags in enumerate(port_day_tags):
        print(f'{jx} {tags}')
        port_day_numtrucks[jx] = len(set(tags))
    print('Completing assessing port entries per truck')
    print(port_day_numtrucks)
    print(f'There were {sum(port_day_numtrucks)} port-day-trucks run this period')

    for cdat in cdata_all:
        type = cdat.Type
        con = cdat.Container
        if type == 'Load In' or type == 'Empty In':
            conid_in.append(cdat.id)
        else:
            cjo = cdat.Jo
            if cjo == 'NAY' or cjo is None:
                conid_nojob.append(cdat.id)
            else:
                mymatch = Interchange.query.filter( (Interchange.Jo == cjo) & (Interchange.id != cdat.id) ).first()
                if mymatch is None:
                    conid_out_only.append(cdat.id)
                    print(f'Container {con} has no Return match')
    print(f'Completed finding containers in range of interest')
    print(f'Found {len(conid_in)} containers returned, {len(conid_out_only)} containers with no return to port, and {len(conid_nojob)} containers with no job reported')
    print(conid_in)


    conid = conid_in + conid_out_only

    # Initiate the running totals
    otot = 0.00
    cashcop = 0.00 # Totals for cash companies
    legitcop = 0.00 # Totals for regular invoicing companies
    nocop = 0.00 # Totals for no company provided should remain zero

    #Initiate the matrixed data
    jobs = []
    nopaycon = []  #Data for containers with no payments made
    nocontainer = []
    issuecon = []
    issuecon2 = []

    pd_amt = [0.00] * num_port_days
    pd_csh = [0.00] * num_port_days
    pd_leg = [0.00] * num_port_days
    pd_non = [0.00] * num_port_days

    pd_col = [0.00] * num_port_days  #Payments collected
    pd_unc = [0.00] * num_port_days  #Payments not collected
    pd_inv = [0.00] * num_port_days  #Payments invoiced

    compare_amt = []

    dict_amounts = {}
    for cat in categories:
        dict_amounts[cat] = [0.00] * num_port_days
    print(dict_amounts)

    def add_to_cat(dict_amounts,itemcode,desc,amt,iport):
        thisvec = dict_amounts['Line Item Total']
        thisvec[iport] += amt
        dict_amounts['Line Item Total'] = thisvec
        for cat in categories:
            if cat == itemcode:
                thisvec = dict_amounts[cat]
                thisvec[iport] += amt
                dict_amounts[cat] = thisvec
                return dict_amounts
        #If got to this point then no cat found
        #Make sure demuurage/per diem not in another cat
        if 'per diem' in desc or 'detention' in desc or 'demurrage' in desc:
            thisvec = dict_amounts['Demurrage']
            thisvec[iport] += amt
            dict_amounts['Demurrage'] = thisvec
        else:
            thisvec = dict_amounts['Other']
            thisvec[iport] += amt
            dict_amounts['Other'] = thisvec
        return dict_amounts


    involist = []
    for cid in conid:
        cdat = Interchange.query.get(cid)
        cdate = cdat.Date
        ccon = cdat.Container
        cjo = cdat.Jo
        istat = None
        if cjo is not None and cjo != 'NAY':
            odat = Orders.query.filter(Orders.Jo == cjo).first()
            if odat is not None:
                try:
                    oamt = float(odat.InvoTotal)
                except:
                    print(f'No invoice total for {odat.Jo}, showing {odat.InvoTotal}, so setting to zero for now')
                    oamt = 0.00
                sh = odat.Shipper
                sid = odat.id
                jo = odat.Jo
                bk = odat.Booking
                con = odat.Container
                d1 = odat.Date
                d2 = odat.Date2

                istat = odat.Istat  # here is key for paid vs not paid
                idat = Invoices.query.filter(Invoices.Jo == cjo).first()
                if idat is not None:
                    iamt = float(idat.Total)
                    tdata = Invoices.query.filter(Invoices.Jo == cjo).all()
                    iport = port_days.index(cdate)
                    ttot = 0.07366
                    involist.append([jo, sh, bk, con, d1, d2, idat.Date, oamt])
                    for tdat in tdata:
                        itemcode = tdat.Service
                        amt = float(tdat.Amount)
                        ttot = ttot + amt
                        desc = tdat.Description
                        desc = desc.lower()
                        dict_amounts = add_to_cat(dict_amounts, itemcode, desc, amt, iport)
                    compare_amt.append([cjo, oamt, iamt, ttot])
                    if int(ttot) == int(iamt):
                        if int(oamt) != int(iamt):
                            oamt = iamt
                            odat.InvoTotal = iamt


                    if abs(oamt-iamt) > .02:
                        issuecon.append([sh, ccon, f'{cdate}', oamt, iamt, 'Invoice amount discrepancy'])
                else:
                    iamt = 0.00
                    issuecon.append([sh, ccon, f'{cdate}', oamt, iamt, 'No invoice found'])

            else:
                rel = cdat.Release
                if rel == 'DP OP' or rel == 'Dry Run Return':
                    sh = 'Noncollect'
                    cjo = 'DP OP'
                    print(f'Drop Pick Return for Operation {ccon}')
                else:
                    print(f'************************No job found for this container {ccon} ***************************')
                    sh = 'No Job'
                    cjo = 'No Job'
                    nocontainer.append([cdate, ccon, cdat.Release])
                    issuecon.append(['Unknown', ccon, f'{cdate}', 0.00, 0.00, f'No job {cjo} found'])
                oamt = 0.00
                istat = 0
        else:
            oamt = 0.00
            rel = cdat.Release
            if rel == 'DP OP' or rel == 'Dry Run Return':
                sh = 'Noncollect'
                cjo = 'DP OP'
                print(f'Drop Pick Return for Operation {ccon}')
                issuecon2.append(['Not Listed', ccon, f'{cdate}', 0.00, 0.00, f'{cjo} operational empty'])
            else:
                print(f'************************No job found for this container {ccon} ***************************')
                sh = 'No Job'
                cjo = 'No Job'
                sid = 0
                jo = 'None'
                nocontainer.append([cdate, ccon, cdat.Release])
                issuecon.append(['Unknown', ccon, f'{cdate}', 0.00, 0.00, f'{cjo} JO listed for this container'])


        iport = port_days.index(cdate)
        pd_amt[iport] += oamt
        otot += oamt

        #Cash vs non-cash collections
        if sh in cashco: pd_csh[iport] += oamt
        elif sh in legitco: pd_leg[iport] += oamt
        else: pd_non[iport] += oamt

        #Paid vs not-paid collections
        if istat is None: istat = 0
        if istat > 1: pd_inv[iport] += oamt
        if istat == 4 or istat == 5 or istat == 8: pd_col[iport] += oamt
        else:
            if sh != 'Noncollect':
                pd_unc[iport] += oamt
                try:
                    nopaycon.append([sid, jo, sh, ccon, f'{cdate}', oamt])
                except:
                    print(f'Could not include nopaycon for container: {ccon}')

        jobs.append([cjo, f'{cdate}', ccon, oamt, otot])

    db.session.commit()
    print(otot)
    print(pd_amt)
    print(dict_amounts)
    print(sum(pd_amt))
    print(f'Average income per port-day = {otot/num_port_days}')
    print(f'Average income per port-day-truck = {otot/sum(port_day_numtrucks)}')

    print('List of all jobs in period')
    #for job in jobs:
        #print(job)

    print('List of jobs without payment')
    #for nopay in nopaycon: print(nopay)

    print('Totals for Specific Dates')
    def get_ytd(cdate, port_days):
        for jx, tdate in enumerate(port_days):
            #print(jx,tdate,cdate)
            if tdate > cdate:
                return jx-1
        return jx

    def get_range(cdate1, cdate2, port_days):
        iport = []
        for jx, tdate in enumerate(port_days):
            if tdate >= cdate1 and tdate <= cdate2:
                iport.append(jx)
            #print(jx,tdate,cdate)
        print(iport)
        if iport != []:
            return min(iport), max(iport)+1
        else:
            return 0, 0

    maxdate = port_days[-1]
    for cdate in keydates:
        if cdate <= maxdate:
            iport = get_ytd(cdate, port_days)
            sub_amt = pd_amt[:iport]
            print(sub_amt)
            print(sum(sub_amt))

    print('Totals by Month')
    ydata = []
    print(pd_amt)
    for month in range(1,ytd_month+1):
        cutoff1 = datetime.date(year, month, 1)
        cutoff2 = last_day_of_month(cutoff1)
        if month == ytd_month: cutoff2 = datetime.date(year, month, ytd_day)
        iport1, iport2 = get_range(cutoff1, cutoff2, port_days)
        if iport1 < 0: iport1 = 0
        if iport2 < 0: iport2 = 0
        print(cutoff1, cutoff2, iport1, iport2)
        sub_amt = pd_amt[iport1:iport2]
        mon_amt = sum(sub_amt)
        print(sub_amt)
        print(sum(sub_amt))
        pd1 = port_days[iport1:iport2]
        npd = len(pd1)
        ptd = port_day_numtrucks[iport1:iport2]
        npdt = sum(ptd)
        if npdt < 1: npdt = 1
        ydata.append([monlist[month-1], sum(pd_amt[iport1:iport2]), sum(pd_csh[iport1:iport2]), sum(pd_leg[iport1:iport2]), sum(pd_col[iport1:iport2]), sum(pd_unc[iport1:iport2]), npd, npdt, mon_amt/npdt])

    print('Totals by Category and Month')
    catdata = []
    for month in range(1,ytd_month+1):
        cutoff1 = datetime.date(year, month, 1)
        cutoff2 = last_day_of_month(cutoff1)
        if month == ytd_month: cutoff2 = datetime.date(year, month, ytd_day)
        iport1, iport2 = get_range(cutoff1, cutoff2, port_days)
        if iport1 < 0: iport1 = 0
        if iport2 < 0: iport2 = 0
        sub_amt = pd_amt[iport1:iport2]
        mon_amt = sum(sub_amt)
        catvec = [monlist[month-1], sum(pd_amt[iport1:iport2])]
        for cat in categories:
            catamtlist = dict_amounts[cat]
            this_sum = sum(catamtlist[iport1:iport2])
            catvec.append(this_sum)
        catdata.append(catvec)
    print(catdata)

    print('Totals by Keydates')
    kdata = []
    print(pd_amt)
    for ix, thisdate in enumerate(keydates):
        if ix == 0:
            cutoff1 = datetime.date(year, 1, 1)
        else:
            cutoff1 = cutoff2 + timedelta(1)
        cutoff2 = thisdate
        if cutoff1 < ytd:
            iport1, iport2 = get_range(cutoff1, cutoff2, port_days)
            if iport1 < 0: iport1 = 0
            if iport2 < 0: iport2 = 0
            print(cutoff1, cutoff2, iport1, iport2)
            sub_amt = pd_amt[iport1:iport2]
            mon_amt = sum(sub_amt)
            print(sub_amt)
            print(sum(sub_amt))
            pd1 = port_days[iport1:iport2]
            npd = len(pd1)
            ptd = port_day_numtrucks[iport1:iport2]
            npdt = sum(ptd)
            if npdt > 0.00:
                thisavg = mon_amt/npdt
            else:
                thisavg = 0.0
            kdata.append([cutoff1, cutoff2, sum(pd_amt[iport1:iport2]), sum(pd_csh[iport1:iport2]), sum(pd_leg[iport1:iport2]), sum(pd_col[iport1:iport2]), sum(pd_unc[iport1:iport2]), npd, npdt, thisavg])

    # Write data to a spreadsheet
    # Give the location of the file spreadsheet to create
    locrec = f'/home/mark/Income/{scac}/{scac}_income_{year}_{ytd_month}_{ytd_day}.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Monthly')
    ws2 = wb.create_sheet(title='Jan')
    #Create column width calculation function
    def column_wide(headers,ydata,other):
        column_widths = []
        for i, cell in enumerate(headers):
            cell = str(cell)
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
        for row in ydata:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for row in other:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        return column_widths

    # Monthly Income Sheet Data Write______________________________________________________________________________
    ws.merge_cells('A1:J1')
    c1 = ws.cell(row=1, column=1, value=f'Report Date:{ytd} -> YTD Workdays={workdays_ytd} of {workdays_yr} total for year')
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(name='Calibri', size=12, bold=True)
    my_gray = Color(rgb='edeef2')
    my_fill = PatternFill(patternType='solid', fgColor=my_gray)
    c1.fill = my_fill

    start_row = 3
    # write the headers
    headers = ['Month', 'Monthly Revenues', 'Cumulative Revenue', ' Cash Jobs ', 'Invoice Jobs', 'Revenue Collected', 'Open Invoices', 'Port Days', 'Port-Truck-Days', 'Income/Port-Truck-Day']
    # define that totals and averages make sense for what columns.  1 we do it, 0 we skip it
    yr_totals = ['YTD Totals:',1,1,1,1,1,1,1,0]
    mon_avg = ['Full-Month Avgs:',1,1,1,1,1,1,1,1]
    proj_totals = ['Projected Year:',1,1,1,0,0,0,0,0]

    for jx, col in enumerate(headers):
        d = ws.cell(row=start_row, column=jx + 1, value=col)
        d.font = Font(name='Calibri',size=10,bold=True)
        d.alignment = Alignment(horizontal='center')

    money = '$#,##0.00'
    dec2 = '#,##0.00'
    dec0 = '#,##0'
    letcol = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    row = start_row
    flist = [0, money, money, money, money, money, dec0, dec0, money]
    for mdata in ydata:
        row = row + 1
        offset = 0
        for ix, col in enumerate(mdata):
            jx = ix + offset + 1
            d = ws.cell(row=row, column=jx, value=col)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            if flist[ix] != 0: d.number_format = flist[ix]
            # Add a formula column for cumlative calcs
            if ix == 1:
                offset = 1
                if row == start_row+1: sumtext = f'= {letcol[ix]}{row}'
                else: sumtext = f'= {letcol[ix]}{row} + {letcol[ix+1]}{row-1}'
                d = ws.cell(row=row, column=ix + offset + 1, value=sumtext)
                d.alignment = Alignment(horizontal='center')
                d.font = Font(name='Calibri', size=10, bold=False)
                d.number_format = money

    sumrow = row + 2
    rowmin = start_row + 1
    rowmax = row

    #Dont include the last month for average if it is not a full month
    if not fullmonth: rowmaxavg = rowmax-1
    else: rowmaxavg = rowmax
    #And do not do the average row if there are no full months


    other1, other2, other3 = [], [], []
    offset = 0
    for ix, col in enumerate(mdata):
        jx = ix+offset+1
        sumtext = f'= SUM({letcol[ix+offset]}{rowmin}:{letcol[ix+offset]}{rowmax})'
        avgtext = f'= AVERAGE({letcol[ix+offset]}{rowmin}:{letcol[ix+offset]}{rowmaxavg})'
        projtext = f'= SUM({letcol[ix+offset]}{rowmin}:{letcol[ix+offset]}{rowmax})*{proj_fraction}'
        #print('this sum text is',sumtext)
        if yr_totals[ix] == 1:
            d = ws.cell(row=sumrow, column=jx, value=sumtext)
            d.number_format = flist[ix]
            other1.append('')
        elif yr_totals[ix] != 0:
            d = ws.cell(row=sumrow, column=jx, value=yr_totals[ix])
            other1.append(yr_totals[ix])
        else: other1.append('')
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)

        if rowmaxavg>=rowmin:
            if mon_avg[ix] == 1:
                e = ws.cell(row=sumrow+1, column=jx, value=avgtext)
                e.number_format = flist[ix]
                other2.append('')
            elif mon_avg[ix] != 0:
                e = ws.cell(row=sumrow+1, column=jx, value=mon_avg[ix])
                other2.append(mon_avg[ix])
            else: other2.append('')
            e.alignment = Alignment(horizontal='center')
            e.font = Font(name='Calibri', size=10, bold=True)

        if proj_totals[ix] == 1:
            f = ws.cell(row=sumrow+3, column=jx, value=projtext)
            f.number_format = flist[ix]
            other3.append('')
        elif proj_totals[ix] != 0:
            f = ws.cell(row=sumrow+3, column=jx, value=proj_totals[ix])
            other3.append(proj_totals[ix])
        else: other3.append('')
        f.alignment = Alignment(horizontal='center')
        f.font = Font(name='Calibri', size=10, bold=True)

        if ix == 1: offset = 1

    other =[other1, other2, other3]
    column_widths = column_wide(headers, ydata, other)
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 0


    # Key Date Income Sheet Data Write______________________________________________________________________________
    ws = wb.create_sheet(title='Weekly')
    ws.merge_cells('A1:k1')
    c1 = ws.cell(row=1, column=1, value=f'Report Date:{ytd} -> YTD Workdays={workdays_ytd} of {workdays_yr} total for year')
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(name='Calibri', size=12, bold=True)
    my_gray = Color(rgb='edeef2')
    my_fill = PatternFill(patternType='solid', fgColor=my_gray)
    c1.fill = my_fill

    start_row = 3
    # write the headers
    headers = ['From', 'To Date', 'Weekly Revenues', 'Cumulative Revenues', ' Cash Jobs ', 'Invoice Jobs', 'Revenue Collected', 'Open Invoices', 'Port Days', 'Port-Truck-Days', 'Income/Port-Truck-Day']
    # define that totals and averages make sense for what columns.  1 we do it, 0 we skip it
    yr_totals = ['YTD Totals:',0,1,1,1,1,1,1,1,0]
    mon_avg = ['Weekly Avgs:',0,1,1,1,1,1,1,1,1]
    proj_totals = ['Projected Year:',0,1,1,1,0,0,0,0,0]

    for jx, col in enumerate(headers):
        d = ws.cell(row=start_row, column=jx + 1, value=col)
        d.font = Font(name='Calibri',size=10,bold=True)
        d.alignment = Alignment(horizontal='center')

    row = start_row
    flist = [0, 0, money, money, money, money, money, dec0, dec0, money]
    for mdata in kdata:
        row = row + 1
        offset = 0
        for ix, col in enumerate(mdata):
            jx = ix + offset + 1
            d = ws.cell(row=row, column=jx, value=col)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            if flist[ix] != 0: d.number_format = flist[ix]
            # Add a formula column for cumlative calcs
            if ix == 2:
                offset = 1
                if row == start_row+1: sumtext = f'= {letcol[ix]}{row}'
                else: sumtext = f'= {letcol[ix]}{row} + {letcol[ix+1]}{row-1}'
                d = ws.cell(row=row, column=ix + offset + 1, value=sumtext)
                d.alignment = Alignment(horizontal='center')
                d.font = Font(name='Calibri', size=10, bold=False)
                d.number_format = money

    sumrow = row + 2
    rowmin = start_row + 1
    rowmax = row

    #Dont include the last month for average if it is not a full month
    if not fullmonth: rowmaxavg = rowmax-1
    else: rowmaxavg = rowmax
    #And do not do the average row if there are no full months

    other1, other2, other3 = [], [], []
    offset = 0
    for ix, col in enumerate(mdata):
        jx = ix + offset + 1
        sumtext = f'= SUM({letcol[ix+offset]}{rowmin}:{letcol[ix+offset]}{rowmax})'
        avgtext = f'= AVERAGE({letcol[ix+offset]}{rowmin}:{letcol[ix+offset]}{rowmaxavg})'
        projtext = f'= SUM({letcol[ix+offset]}{rowmin}:{letcol[ix+offset]}{rowmax})*{proj_fraction}'
        #print('this sum text is',sumtext)

        if yr_totals[ix] == 1:
            d = ws.cell(row=sumrow, column=jx, value=sumtext)
            d.number_format = flist[ix]
            other1.append('')
        elif yr_totals[ix] != 0:
            d = ws.cell(row=sumrow, column=jx, value=yr_totals[ix])
            other1.append(yr_totals[ix])
        else: other1.append('')
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)

        if rowmaxavg>=rowmin:
            if mon_avg[ix] == 1:
                e = ws.cell(row=sumrow+1, column=jx, value=avgtext)
                e.number_format = flist[ix]
                other2.append('')
            elif mon_avg[ix] != 0:
                e = ws.cell(row=sumrow+1, column=jx, value=mon_avg[ix])
                other2.append(mon_avg[ix])
            else: other2.append('')
            e.alignment = Alignment(horizontal='center')
            e.font = Font(name='Calibri', size=10, bold=True)

        if proj_totals[ix] == 1:
            f = ws.cell(row=sumrow+3, column=jx, value=projtext)
            f.number_format = flist[ix]
            other3.append('')
        elif proj_totals[ix] != 0:
            f = ws.cell(row=sumrow+3, column=jx, value=proj_totals[ix])
            other3.append(proj_totals[ix])
        else: other3.append('')
        f.alignment = Alignment(horizontal='center')
        f.font = Font(name='Calibri', size=10, bold=True)
        if ix == 2: offset = 1

    other =[other1, other2, other3]
    column_widths = column_wide(headers, kdata, other)
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 0

    ########################################################################################################
    ########################################################################################################
    # Category Income Sheet Data Write______________________________________________________________________________
    ws = wb.create_sheet(title='Categories')
    ws.merge_cells('A1:L1')
    c1 = ws.cell(row=1, column=1, value=f'Report Date:{ytd} -> YTD Workdays={workdays_ytd} of {workdays_yr} total for year')
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(name='Calibri', size=12, bold=True)
    my_gray = Color(rgb='edeef2')
    my_fill = PatternFill(patternType='solid', fgColor=my_gray)
    c1.fill = my_fill
    money = '$#,##0.00'
    dec2 = '#,##0.00'
    dec0 = '#,##0'
    start_row = 3
    # write the headers
    headers = ['Month', 'Monthly Revenues']
    flist = [0, money]
    yr_totals = ['YTD Totals:',1]
    for cat in categories:
        headers.append(cat)
        flist.append(money)
        yr_totals.append(1)
    # define that totals and averages make sense for what columns.  1 we do it, 0 we skip it

    for jx, col in enumerate(headers):
        d = ws.cell(row=start_row, column=jx + 1, value=col)
        d.font = Font(name='Calibri',size=10,bold=True)
        d.alignment = Alignment(horizontal='center')

    row = start_row

    for mdata in catdata:
        row = row + 1
        for ix, col in enumerate(mdata):
            d = ws.cell(row=row, column=ix + 1, value=col)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            if flist[ix] != 0: d.number_format = flist[ix]
        checksumtext = f'= SUM(D{row}:K{row})'
        d = ws.cell(row=row, column=12, value=checksumtext)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money
    sumrow = row + 2
    rowmin = start_row + 1
    rowmax = row

    #Dont include the last month for average if it is not a full month
    if not fullmonth: rowmaxavg = rowmax-1
    else: rowmaxavg = rowmax
    #And do not do the average row if there are no full months

    letcol = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
    other1, other2, other3 = [], [], []
    for ix, col in enumerate(mdata):
        sumtext = f'= SUM({letcol[ix]}{rowmin}:{letcol[ix]}{rowmax})'
        #print('this sum text is',sumtext)

        if yr_totals[ix] == 1:
            d = ws.cell(row=sumrow, column=ix + 1, value=sumtext)
            d.number_format = flist[ix]
            other1.append('')
        elif yr_totals[ix] != 0:
            d = ws.cell(row=sumrow, column=ix + 1, value=yr_totals[ix])
            other1.append(yr_totals[ix])
        else: other1.append('')
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)

    other = [other1]
    column_widths = column_wide(headers, catdata, other)
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 4
    ########################################################################################################

    #Missing Containers Write

    ws = wb.create_sheet(title='Missing Containers')
    row = 1
    # write the headers
    headers = ['Date', 'Container', 'Release']

    for jx, col in enumerate(headers):
        d = ws.cell(row=1, column=jx + 1, value=col)
        d.font = Font(name='Calibri', size=10, bold=True)
        d.alignment = Alignment(horizontal='center')
    for noco in nocontainer:
        row = row + 1
        for ix, col in enumerate(noco):
            d = ws.cell(row=row, column=ix + 1, value=col)
            d.font = Font(name='Calibri', size=10, bold=False)
            d.alignment = Alignment(horizontal='center')

    column_widths = column_wide(headers, nocontainer,[])
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 3

    ws = wb.create_sheet(title='OpenBal')
    row = 1
    # write the headers
    headers = ['ID', 'JO', 'Company', 'Container', 'Date', 'Amount']
    flist = [0, 0, 0, 0, 0, money]
    for jx, col in enumerate(headers):
        d = ws.cell(row=1, column=jx + 1, value=col)
        d.font = Font(name='Calibri', size=10, bold=True)
        d.alignment = Alignment(horizontal='center')
    for nopay in nopaycon:
        row = row + 1
        for ix, col in enumerate(nopay):
            d = ws.cell(row=row, column=ix + 1, value=col)
            d.font = Font(name='Calibri', size=10, bold=False)
            d.alignment = Alignment(horizontal='center')
            if flist[ix] != 0: d.number_format = flist[ix]

    column_widths = column_wide(headers, nopaycon, [])
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 3

    # Now create a sheet for every company and provide more about those unpaid jobs:
    ucos = [] #unique companies with open balances
    for nopay in nopaycon:
        company = nopay[2]
        if company not in ucos:
            if company != 'No Job':
                ucos.append(company)




    if 1 == 1:
        ws = wb.create_sheet(title='Issues')
        row = 1
        flist = [0, 0, 0, money, money, 0]
        # write the headers
        headers = ['Company', 'Container', 'Date', 'Amount InvoTot', 'Amount Invoiced', 'Comment']
        for jx, col in enumerate(headers):
            d = ws.cell(row=1, column=jx + 1, value=col)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=True)

        for iss in issuecon:
            row = row + 1
            for ix, col in enumerate(iss):
                d = ws.cell(row=row, column=ix + 1, value=col)
                d.alignment = Alignment(horizontal='center')
                d.font = Font(name='Calibri', size=10, bold=False)
                if flist[ix] != 0: d.number_format = flist[ix]

    column_widths = column_wide(headers, issuecon, [])
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 3

    if 1 == 1:
        ws = wb.create_sheet(title='Operational Moves')
        row = 1
        flist = [0, 0, 0, money, money, 0]
        # write the headers
        headers = ['Company', 'Container', 'Date', 'Amount InvoTot', 'Amount Invoiced', 'Comment']
        for jx, col in enumerate(headers):
            d = ws.cell(row=1, column=jx + 1, value=col)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=True)

        for iss in issuecon2:
            row = row + 1
            for ix, col in enumerate(iss):
                d = ws.cell(row=row, column=ix + 1, value=col)
                d.alignment = Alignment(horizontal='center')
                d.font = Font(name='Calibri', size=10, bold=False)
                if flist[ix] != 0: d.number_format = flist[ix]

    column_widths = column_wide(headers, issuecon2, [])
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 3

    for uco in ucos:
        print('uco=', uco)
        nco = uco.replace('/','')
        nco = nco.split()
        if len(nco) == 1: bco = f'{nco[0]}'
        else: bco = f'{nco[0]} {nco[1]}'
        ws = wb.create_sheet(title=bco)

        #
        odata = Orders.query.filter((Orders.Shipper == uco) & (Orders.Istat<4)).all()

        row = 1
        # write the headers
        headers = ['ID', 'JO', 'Company', 'Booking', 'Container', 'Date Out', 'Date In', 'Invoice Date', 'Invoice Amount']
        flist = [0, 0, 0, 0, 0, 0, 0, 0, money]

        for jx, col in enumerate(headers):
            d = ws.cell(row=1, column=jx + 1, value=col)
            d.font = Font(name='Calibri', size=10, bold=True)
            d.alignment = Alignment(horizontal='center')

        #for nopay in nopaycon:
        for odat in odata:
            if odat.Istat >= 2:
                thisbk = odat.Booking
                idat = Invoices.query.filter(Invoices.Jo == odat.Jo).first()
                if idat is not None: idate = idat.Date
                else: idate = 'No Invoice'
                diffbk = 0
                intdata = Interchange.query.filter(Interchange.Jo == jo).all()
                intlen = len(intdata)
                if intlen == 2:
                    int1 = intdata[0]
                    int2 = intdata[1]
                    bk1 = int1.Release
                    bk2 = int2.Release
                    #print('Booking Pair:', bk1, bk2)
                    if bk1 != bk2:
                        diffbk = 1
                        thisbk = f'{bk1}:{bk2}'
                row = row + 1
                try:
                    amt = float(odat.InvoTotal)
                except:
                    amt = 0.00

                rowdat = [odat.id, odat.Jo, odat.Shipper, thisbk, odat.Container, odat.Date, odat.Date2, idate, amt]
                for ix, col in enumerate(rowdat):
                    d = ws.cell(row=row, column=ix + 1, value=col)
                    d.font = Font(name='Calibri', size=10, bold=False)
                    d.alignment = Alignment(horizontal='center')
                    if flist[ix] != 0: d.number_format = flist[ix]

        sumtext = f'= SUM(I2:I{row})'
        d = ws.cell(row=row+1, column=9, value=sumtext)
        d.font = Font(name='Calibri', size=10, bold=True)
        d.alignment = Alignment(horizontal='center')
        d.number_format = money

        column_widths = column_wide(headers, nopaycon, [])
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 3

    row = 1
    # write the headers
    headers = ['JO', 'Company', 'Booking', 'Container', 'Date Out', 'Date In', 'Invoice Date', 'Invoice Amount']
    flist = [0, 0, 0, 0, 0, 0, 0, money]

    for jx, col in enumerate(headers):
        d = ws.cell(row=1, column=jx + 1, value=col)
        d.font = Font(name='Calibri', size=10, bold=True)
        d.alignment = Alignment(horizontal='center')
    for invo in involist:
        jo = invo[0]
        sh = invo[1]
        amt = invo[2]
        date = invo[3]
        print(f'{jo} {sh} {amt} {date}')
        row += 1
        for jx, item in enumerate(invo):
            d = ws2.cell(row=row, column=jx+1, value=item)
            d.font = Font(name='Calibri', size=10, bold=False)
            d.alignment = Alignment(horizontal='center')

    std=wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)
    wb.save(locrec)

    for comp in compare_amt:
        if int(comp[1]) != int(comp[2]) or int(comp[1]) != int(comp[3]):
            print(comp)

else:
    print(f'Could not open database {scac}')


tunnel.stop()