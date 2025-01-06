import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
from utils import d2f, d2s, hasinput
import datetime
import socket
from utils import getpaths
from datetime import timedelta
from pyvirtualdisplay import Display
from selenium import webdriver
import time

#Handle the input arguments from script file
try:
    scac = sys.argv[1]
except:
    print('Must have at least one argument...FELA or OSLM or NEVO')
    scac = 'fela'

scac = scac.upper()
nt = 'remote'

if scac == 'OSLM' or scac == 'FELA' or scac == 'NEVO':
    print(f'Running FFF_dray_Chassis.py for {scac}')
    host_name = socket.gethostname()
    print("Host Name:", host_name)
    dropbox_path = getpaths(host_name, 'dropbox')
    sys_path = getpaths(host_name, 'system')
    sys.path.append(sys_path) #So we can import CCC_system_setup from full path

    os.environ['SCAC'] = scac
    os.environ['PURPOSE'] = 'script'
    os.environ['MACHINE'] = host_name
    os.environ['TUNNEL'] = nt

    from remote_db_connect import tunnel, db
    from models8 import Orders, Interchange, Invoices

else:
    scac = 'nogo'
    print('The argument must be FELA or OSLM or NEVO')
    quit()

if scac != 'nogo':

    today = datetime.datetime.today()
    year = today.year
    yearbegin = datetime.date(year, 1, 1)

    today_str = today.strftime("%m/%d/%Y")
    money = '$#,##0.00'

    #Define paths to files:
    trac_path = f'{dropbox_path}/Dray/{scac}/Chassis/TRAC'
    dcli_path = f'{dropbox_path}/Dray/{scac}/Chassis/DCLI'
    summ_path = f'{dropbox_path}/Dray/{scac}_SummaryUpdate.xlsx'

    print(f'the summ_path is {summ_path}')

    #Define border styles:
    thickbl = Side(border_style='thick', color = '050505')
    bottom = Side(border_style='thick', color = '050505')
    left = Side(border_style='thick', color = '050505')
    right = Side(border_style='thick', color = '050505')
    expfill = PatternFill(fgColor="8AD4ED", fill_type = "solid")
    impfill = PatternFill(fgColor="EDBFA4", fill_type = "solid")

    printif = 1

    #######Chassis Invoice Checker############

    #######TRACTRACTRACTRACTRACTRACTRACTRACTRACTRACTRAC#####################################################################
    df = pd.read_excel(summ_path, sheet_name='TRAC Invoices')
    df2 = pd.read_excel(summ_path, sheet_name='TRAC Reconciliation')
    total_row_invo = df.shape[0]
    total_row_recon = df2.shape[0]
    df_start_row = total_row_invo + 1
    df2_start_row = total_row_recon + 2
    print('Start Rows',df_start_row,df2_start_row)
    #sh2 = pd.read_excel(summ_path, sheet_name='TRAC Reconciliation')

    invo_old = []
    invo_new = []
    errors = 0

    for i in df.index:
        invo = df['Invoice'][i]
        invo_old.append(invo)

    print(f'The invoices already recorded are: {invo_old}')

    dlist = os.listdir(trac_path)
    print(f'The list of all invoice in folder is {dlist}')

    oxl = openpyxl.load_workbook(summ_path)
    ws = oxl['TRAC Reconciliation']
    wv = oxl['TRAC Invoices']

    for tfile in dlist:
        invonum = tfile.replace('Invoice_','').replace('.xlsx','')
        invonum = int(invonum)
        print(invonum)
        if invonum in invo_old:
            print(f'{tfile} is old')
        else:
            print(f'{tfile} is new')
            invo_new.append(tfile)

    print(f'The new TRAC invoices to be added are: {invo_new}')
    thisrow = df_start_row
    nextrow = df2_start_row
    # Reconcile the new invoices
    for tfile in invo_new:
        amt_total = 0.00
        amt_coll = 0.00
        ptfile = f'{trac_path}/{tfile}'
        shname= tfile.replace('.xlsx', '')
        invonum = int(shname.replace('Invoice_',''))

        dp = pd.read_excel(ptfile, sheet_name=shname)
        for i in dp.index:
            chassis = str(dp['CHASSIS ID'][i])
            container = str(dp['DROP OFF CONTAINER'][i])
            container = container.replace('-','')
            if container == 'nan':
                container = 'None'
            if container == 'None':
                container = str(dp['PICK UP CONTAINER'][i])
                container = container.replace('-', '')

            days = int(dp['BILLING UNITS'][i])
            rate = float(dp['BILLING RATE'][i])
            subtotal = rate*days
            amt_total = amt_total + subtotal
            if days > 0:
                success = 0
                trys = 0
                while success == 0 and trys < 20:
                    if container != 'None':
                        try:
                            kdata = Interchange.query.filter(Interchange.Container == container).all()
                            success = 1
                        except:
                            print('Could not open tunnel 1')
                            trys = trys + 1

                nticks = len(kdata)
                if nticks > 2:
                    print(f'Error working on container {container} or chassis {chassis}')
                    print(f'nticks is {nticks}')
                    errors = errors+1
                if nticks == 2:

                    nextrow = nextrow + 1
                    k1 = kdata[0]
                    k2 = kdata[1]
                    mydays = k2.Date - k1.Date
                    mydays = mydays.days + 1
                    jo = k1.Jo
                    idat = Invoices.query.filter((Invoices.Jo==jo) & (Invoices.Service == 'Chassis Fees')).first()
                    if idat is not None:
                        invodol = idat.Amount
                    else:
                        invodol = 0
                    amt_coll = amt_coll + float(invodol)

                    print(f'Working container {container} from {k1.Date} to {k2.Date}')
                    items = [today_str, invonum, container, chassis, days, d2f(subtotal), k1.Container, k1.Chassis,
                             k1.Date, k2.Date, mydays, k1.Company, d2f(invodol)]

                    fstyle = [0]*13
                    fstyle[5] = 'money'
                    fstyle[12] = 'money'
                    for jx, item in enumerate(items):
                        d=ws.cell(row=nextrow, column=jx+1, value=item)
                        d.alignment = Alignment(horizontal='center')
                        d.font = Font(name='Calibri', size=10, bold=False)
                        if fstyle[jx] == 'money': d.number_format = money

                if nticks == 1:
                    nextrow = nextrow + 1
                    k1 = kdata[0]
                    jo = k1.Jo
                    idat = Invoices.query.filter((Invoices.Jo==jo) & (Invoices.Service == 'Chassis Fees')).first()
                    if idat is not None:
                        invodol = idat.Amount
                    else:
                        invodol = 0
                    amt_coll = amt_coll + float(invodol)
                    items = [today_str, invonum, container, chassis, days, d2f(subtotal), k1.Container, k1.Chassis, k1.Date, k1.Date, '', k1.Company, d2f(invodol)]
                    fstyle = [0]*13
                    fstyle[5] = 'money'
                    fstyle[12] = 'money'
                    print(f'Working container {container} one ticket on {k1.Date}')
                    for jx, item in enumerate(items):
                        d=ws.cell(row=nextrow, column=jx+1, value=item)
                        d.alignment = Alignment(horizontal='center')
                        d.font = Font(name='Calibri', size=10, bold=False)
                        if fstyle[jx] == 'money': d.number_format = money

                if nticks == 0:
                    nextrow = nextrow + 1
                    items = [today_str, invonum, container, chassis, days, d2f(subtotal), '', '', '', '', '', 'Unknown', '']
                    fstyle = [0]*13
                    fstyle[5] = 'money'

                    for jx, item in enumerate(items):
                        d=ws.cell(row=nextrow, column=jx+1, value=item)
                        d.alignment = Alignment(horizontal='center')
                        d.font = Font(name='Calibri', size=10, bold=False)
                        if fstyle[jx] == 'money': d.number_format = money


        thisrow = thisrow + 1
        nextrow = nextrow + 2
        items = [today_str, invonum, '', d2f(amt_total), d2f(amt_coll)]
        fstyle = [0]*5
        fstyle[3] = 'money'
        fstyle[4] = 'money'
        for jx, item in enumerate(items):
            d = wv.cell(row=thisrow, column=jx + 1, value=item)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            if fstyle[jx] == 'money': d.number_format = money


    oxl.save(summ_path)




    #######DCLIDCLIDCLIDCLIDCLIDCLIDCLIDCLIDCLIDCLIDCLI#####################################################################
    df = pd.read_excel(summ_path, sheet_name='DCLI Invoices')
    df2 = pd.read_excel(summ_path, sheet_name='DCLI Reconciliation')
    total_row_invo = df.shape[0]
    total_row_recon = df2.shape[0]
    df_start_row = total_row_invo + 1
    df2_start_row = total_row_recon + 2
    print('Start Rows',df_start_row,df2_start_row)

    invo_old = []
    invo_new = []

    for i in df.index:
        invo = df['Invoice'][i]
        invo_old.append(invo)

    print('DCLI Old Invoices', invo_old)

    dlist = os.listdir(dcli_path)
    print('DCLI All invoice files in folder:', dlist)

    oxl = openpyxl.load_workbook(summ_path)
    ws = oxl['DCLI Reconciliation']
    wv = oxl['DCLI Invoices']

    for tfile in dlist:
        print(f'DCLI filename is: {tfile}')
        invonum = tfile.replace('invoice_audit','').replace('.xls','')
        print(invonum)
        if invonum in invo_old:
            print(f'{tfile} is old')
        else:
            print(f'{tfile} is new')
            invo_new.append(tfile)

    print(f'The new DCLI invoices to be added are: {invo_new}')
    thisrow = df_start_row
    nextrow = df2_start_row
    # Reconcile the new invoices
    for tfile in invo_new:
        amt_total = 0.00
        amt_coll = 0.00
        ptfile = f'{dcli_path}/{tfile}'
        shname= tfile.replace('.xls', '')
        invonum = shname.replace('invoice_audit','')
        dp = pd.read_excel(ptfile, sheet_name='invoice_audit', header=1)
        for i in dp.index:
            chassis = str(dp['Chassis In'][i])
            container = str(dp['Container In'][i])
            container = container.replace('-','')
            if container == 'nan':
                container = 'None'
            try:
                days = int(dp['Tier 1 Days'][i])
            except:
                days = int(dp['Days'][i])
            try:
                rate = float(dp['Tier 1 Rate'][i])
            except:
                rate = float(dp['Rate'][i])
            try:
                tax = float(dp['Tax Amount'][i])
            except:
                tax = 0.00
            subtotal = rate*days + tax
            amt_total = amt_total + subtotal
            print(f'Results for container {container} with chassis {chassis} and days {days} with total {subtotal}')
            if days > 0:
                success = 0
                trys = 0
                while success == 0 and trys < 20:
                    if container != 'None':
                        try:
                            kdata = Interchange.query.filter(Interchange.Container == container).all()
                            success = 1
                        except:
                            print('Could not open tunnel 1')
                            trys = trys + 1
                    else:
                        try:
                            kdata = Interchange.query.filter(Interchange.Chassis == chassis).all()
                            success = 1
                        except:
                            print('Could not open tunnel 2')
                            trys = trys + 1
                nticks = len(kdata)
                print(nticks)
                if nticks == 2:

                    k1 = kdata[0]
                    k2 = kdata[1]
                    mydays = k2.Date - k1.Date
                    mydays = mydays.days + 1
                    jo = k1.Jo
                    idat = Invoices.query.filter((Invoices.Jo==jo) & (Invoices.Service == 'Chassis Fees')).first()
                    if idat is not None:
                        invodol = idat.Amount
                    else:
                        invodol = 0
                    amt_coll = amt_coll + float(invodol)
                    ws.cell(row=nextrow, column=1, value=invonum)
                    ws.cell(row=nextrow, column=2, value=container)
                    ws.cell(row=nextrow, column=3, value=chassis)
                    ws.cell(row=nextrow, column=4, value=days)
                    ws.cell(row=nextrow, column=5, value=d2f(subtotal))
                    ws.cell(row=nextrow, column=6, value=k1.Container)
                    ws.cell(row=nextrow, column=7, value=k1.Chassis)
                    ws.cell(row=nextrow, column=8, value=k1.Date)
                    ws.cell(row=nextrow, column=9, value=k2.Date)
                    ws.cell(row=nextrow, column=10, value=mydays)
                    ws.cell(row=nextrow, column=11, value=k1.Company)
                    ws.cell(row=nextrow, column=12, value=d2f(invodol))
                    nextrow = nextrow + 1

                if nticks == 1:

                    k1 = kdata[0]
                    jo = k1.Jo
                    idat = Invoices.query.filter((Invoices.Jo==jo) & (Invoices.Service == 'Chassis Fees')).first()
                    if idat is not None:
                        invodol = idat.Amount
                    else:
                        invodol = 0
                    amt_coll = amt_coll + float(invodol)

                    ws.cell(row=nextrow, column=1, value=invonum)
                    ws.cell(row=nextrow, column=2, value=container)
                    ws.cell(row=nextrow, column=3, value=chassis)
                    ws.cell(row=nextrow, column=4, value=days)
                    ws.cell(row=nextrow, column=5, value=d2f(subtotal))
                    ws.cell(row=nextrow, column=6, value=k1.Container)
                    ws.cell(row=nextrow, column=7, value=k1.Chassis)
                    ws.cell(row=nextrow, column=8, value=k1.Date)
                    ws.cell(row=nextrow, column=11, value=k1.Company)
                    ws.cell(row=nextrow, column=12, value=d2f(invodol))
                    nextrow = nextrow + 1

                if nticks == 0:

                    invodol = 0.00
                    amt_coll = 0.00

                    ws.cell(row=nextrow, column=1, value=invonum)
                    ws.cell(row=nextrow, column=2, value=container)
                    ws.cell(row=nextrow, column=3, value=chassis)
                    ws.cell(row=nextrow, column=4, value=days)
                    ws.cell(row=nextrow, column=5, value=d2f(subtotal))
                    ws.cell(row=nextrow, column=6, value='None')
                    ws.cell(row=nextrow, column=7, value='None')
                    ws.cell(row=nextrow, column=12, value=d2f(invodol))
                    nextrow = nextrow + 1

                if nticks == 3:

                    k1 = kdata[0]
                    k2 = kdata[1]
                    k3 = kdata[2]
                    lowboy = min(k1.Date, k2.Date, k3.Date)
                    hiboy = max(k1.Date, k2.Date, k3.Date)
                    mydays = hiboy - lowboy
                    mydays = mydays.days + 1
                    jo = k1.Jo
                    idat = Invoices.query.filter((Invoices.Jo == jo) & (Invoices.Service == 'Chassis Fees')).first()
                    if idat is not None:
                        invodol = idat.Amount
                    else:
                        invodol = 0
                    amt_coll = amt_coll + float(invodol)
                    ws.cell(row=nextrow, column=1, value=invonum)
                    ws.cell(row=nextrow, column=2, value=container)
                    ws.cell(row=nextrow, column=3, value=chassis)
                    ws.cell(row=nextrow, column=4, value=days)
                    ws.cell(row=nextrow, column=5, value=d2f(subtotal))
                    ws.cell(row=nextrow, column=6, value=k1.Container)
                    ws.cell(row=nextrow, column=7, value=k1.Chassis)
                    ws.cell(row=nextrow, column=8, value=lowboy)
                    ws.cell(row=nextrow, column=9, value=hiboy)
                    ws.cell(row=nextrow, column=10, value=mydays)
                    ws.cell(row=nextrow, column=11, value=k1.Company)
                    ws.cell(row=nextrow, column=12, value=d2f(invodol))
                    nextrow = nextrow + 1


        thisrow = thisrow + 1
        nextrow = nextrow + 1
        wv.cell(row=thisrow, column=1, value=invonum)
        wv.cell(row=thisrow, column=2, value=today_str)
        wv.cell(row=thisrow, column=5, value=d2f(amt_total))
        wv.cell(row=thisrow, column=6, value=d2f(amt_coll))

        ###################### Accrual Method Chassis Rentals by Date ##########################################
    odata = Orders.query.filter(Orders.Date2 >= yearbegin).all()
    dclis = ['MSCZ', 'APMZ', 'DCLZ']
    tracs = ['MRTZ', 'METZ', 'MAEU', 'TSXZ']
    ws2 = oxl.create_sheet(title='Chassis Rentals')
    rowdat = []
    oslmdat = []
    row = 2

    for odat in odata:
        out = odat.Date
        oin = odat.Date2
        try:
            days = oin - out
            days = days.days + 1
        except:
            days = 0
        ch = odat.Chassis
        if ch is not None:
            if len(ch) >= 4:
                ffch = ch[:4]
            else:
                ffch = ch
        else:
            ffch = ''
        if ffch in dclis:
            print("DCLI",out,oin,days,ch,ffch)
            tot = days*30.78
            rowdat.append(['DCLI', out, oin, days, ch, ffch, 30.78, tot])
        elif ffch in tracs:
            print("TRAC",out,oin,days,ch,ffch)
            tot = days*29.75
            rowdat.append(['TRAC', out, oin, days, ch, ffch, 29.75, tot])
        else:
            tot = days * 35.00
            print("OSLM",out,oin,days,ch,ffch)
            oslmdat.append(['OSLM', out, oin, days, ch, ffch, 35.00, tot])

    for jx,row in enumerate(rowdat):
        for ix, col in enumerate(row):
            ws2.cell(row=jx+2, column=ix+1, value=col)



    oxl.save(summ_path)
    tunnel.stop()