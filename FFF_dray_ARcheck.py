import os
import sys
import socket
from utils import getpaths

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
from utils import d2s, stripper, hasinput
import datetime
from datetime import timedelta


today = datetime.datetime.today()
today_str = today.strftime("%m/%d/%Y")
d = today.strftime("%B %d, %Y")
#Calc days back desered to go back to last date of year two years prior
tyear = today.year - 1
last_day_back = datetime.date(tyear, 1, 1)
daysback = today.date() - last_day_back
days_far_back = daysback.days
cutoff = datetime.datetime.now() - timedelta(365)
cutoff = cutoff.date()
over30 = datetime.datetime.now() - timedelta(30)
over30 = over30.date()
todaydate = today.date()

#Handle the input arguments from script file
try:
    scac = sys.argv[1]
    nt = 'remote'
except:
    print('Must have at least one argument...FELA or OSLM or NEVO')
    scac = 'FELA'
    nt = 'remote'

scac = scac.upper()

if scac == 'OSLM' or scac == 'FELA' or scac == 'NEVO':

    print(f'Running FFF_dray_ARcheck for {scac}')
    host_name = socket.gethostname()
    print("Host Name:", host_name)
    dropbox_path = getpaths(host_name, 'dropbox')
    ar_path = f'{dropbox_path}/Dray/{scac}_AR_Report.xlsx'
    sys_path = getpaths(host_name, 'system')
    sys.path.append(sys_path) #So we can import CCC_system_setup from full path

    os.environ['SCAC'] = scac
    os.environ['PURPOSE'] = 'script'
    os.environ['MACHINE'] = host_name
    os.environ['TUNNEL'] = nt

    from remote_db_connect import db
    if nt == 'remote': from remote_db_connect import tunnel
    from models8 import Orders, Interchange, Invoices

else:
    scac = 'nogo'
    print('The argument must be FELA or OSLM or NEVO')
    quit()

if scac != 'nogo':
    wb = openpyxl.load_workbook(ar_path, data_only=True)

    shtlist = wb.sheetnames
    for sht in shtlist:
        if sht != 'AR Notes':
            std = wb.get_sheet_by_name(sht)
            wb.remove_sheet(std)

    sheet_str = today_str.replace('/',' ')
    newsheet = f'AR Summary {sheet_str}'
    dfs= wb.create_sheet(newsheet)
    newsheet = f'AR by Date {sheet_str}'
    dfc= wb.create_sheet(newsheet)

    #formats for writing to excel
    money = '$#,##0.00'
    dec2 = '#,##0.00'
    dec0 = '#,##0'

    #Create column width calculation function
    def column_wide(headers,ydata):
        column_widths = []
        for cell in enumerate(headers):
            cell = str(cell)
            column_widths.append(len(cell))
        for row in ydata:
            for i, cell in enumerate(row):
                test = len(str(cell))
                column_widths[i] = max(test,column_widths[i])
        return column_widths

    #Write the header for AR by Date Sheet
    row = 1
    ydata = []
    hdrs = ['SID','JO','Company', 'Order', 'Release','Booking In','Container','Date Out','Date In','Date Invoiced', 'Amount','Invoice Filename', 'Comment ID']
    for col, hdr in enumerate(hdrs):
        d = dfc.cell(row=row, column=col + 1, value=hdr)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
    success = 0
    trys = 0
    while success == 0 and trys < 20:
        try:
            odata = Orders.query.filter((Orders.Istat > 0) & (Orders.Istat < 5) & (Orders.Date3 > cutoff)).all()
            success = 1
        except:
            print(f'Could not open tunnel on try {trys}')
            trys = trys + 1

    if success == 1:
        #Make a list of each unique company with invoices in the past year:
        comps = []
        for odat in odata:
            comp = odat.Shipper
            if comp not in comps: comps.append(comp)
        print(f'Completed getting unique companies with invoices in past year')

        def defset(str):
            if hasinput(str): return str
            else: return ''

        for gdat in odata:
            jo = gdat.Jo
            odr = defset(gdat.Order)
            bk1 = defset(gdat.Booking)
            bk2 = defset(gdat.BOL)
            con = defset(gdat.Container)
            dt1 = gdat.Date
            dt2 = gdat.Date2
            invo = defset(gdat.Invoice)
            links = defset(gdat.Links)
            if dt1 is not None and dt2 is not None:
                idat = Invoices.query.filter(Invoices.Jo == jo).first()
                if idat is not None:
                    dti = idat.Date
                    amt = float(idat.Total)

                    #Make sure the invototal matches the invoice:
                    invocheck = defset(gdat.InvoTotal)
                    test = d2s(amt)
                    if test != invocheck:
                        print(f'Fixing Orders InvoTotal to match the Invoice Total for jo {jo}')
                        gdat.InvoTotal = test
                        db.session.commit()

                    newblock = [gdat.id, gdat.Jo, gdat.Shipper, odr, bk1, bk2, con, dt1, dt2, dti, amt, invo, links]
                    ydata.append(newblock)
                else:
                    dti = 'None'
                    amt = 'No Invoice'
                    print(f'No invoice found for {gdat.id} with JO: {jo}')
            else:
                print(f'Do not have both and out and in date for {gdat.id} with JO: {jo}....need to fix')

        ydata.sort(key=lambda row: (row[9], row[12]))

        for ydat in ydata:
            row = row + 1
            dti = ydat[9]

            for col, item in enumerate(ydat):
                d = dfc.cell(row=row, column=col + 1, value=item)
                d.alignment = Alignment(horizontal='center')
                if dti < over30:
                    d.font = Font(name='Calibri', size=10, bold=False)
                else:
                    d.font = Font(name='Calibri', size=10, bold=False)

                if col == 10: d.number_format = money

        column_widths = column_wide(hdrs, ydata)
        for i, column_width in enumerate(column_widths):
            dfc.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

        #dfc.auto_filter.ref = f'A1:M{row}'
        #dfc.auto_filter.add_filter_column(8, ['Inactive'])
        #dfc.auto_filter.add_sort_condition(f'I2:I{row}')
        sumover_all = 0.00
        sumunder_all = 0.00
        zdata = []
        s_hdrs = ['Company', 'Over 30 Days', 'Under 30 Days', 'Total Unpaid Invoices']
        desc = f'Report last updated on {todaydate}'
        #Kill all then entries and add them each in

        for col, hdr in enumerate(s_hdrs):
            d = dfs.cell(row=1, column=col + 2, value=hdr)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=True)

        #Write tabs for each company
        exclude_these = ['FEL Ocean Division', 'First Eagle Logistics', 'One Stop Logistics', 'Nello Enterprise LLC', 'Jays Auto Service']
        for comp in comps:
            if not any(comp in x for x in exclude_these):
                tab = comp.split()
                if len(tab)>1: newtab = f'{tab[0]} {tab[1]}'
                else: newtab = f'{tab[0]}'
                newtab = newtab.replace(',','')
                print(newtab)

                # Set up variables to sum the information for each company
                sumover = 0.0
                sumunder = 0.0

                ws = wb.create_sheet(title=newtab)
                row = 1
                # write the headers
                hdrs = ['SID', 'JO', 'Company', 'Order', 'Release', 'Booking In', 'Container', 'Date Out', 'Date In',
                        'Date Invoiced', 'Amount', 'Invoice Filename', 'Comment ID']
                for col, hdr in enumerate(hdrs):
                    d = ws.cell(row=row, column=col + 1, value=hdr)
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name='Calibri', size=10, bold=True)

                for ydat in ydata:
                    shp = ydat[2]
                    if comp == shp:
                        row = row + 1
                        dti = ydat[9]
                        if dti < over30: sumover += ydat[10]
                        else: sumunder += ydat[10]

                        for col, item in enumerate(ydat):
                            d = ws.cell(row=row, column=col + 1, value=item)
                            d.alignment = Alignment(horizontal='center')
                            if dti < over30:
                                d.font = Font(name='Calibri', size=10, bold=False)
                            else:
                                d.font = Font(name='Calibri', size=10, bold=False)
                            if col == 10: d.number_format = money

                for i, column_width in enumerate(column_widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

                #Write the Summary Data for Company to Summary Page
                zdata.append([comp,sumover,sumunder,sumover+sumunder])
                sumunder_all += sumunder
                sumover_all += sumover

        row_s = 1
        zdata.sort(key=lambda row: row[1], reverse=True)
        for zdat in zdata:
            comp, so, su, stot = zdat
            compu = comp.upper()

            row_s += 1

            d = dfs.cell(row=row_s, column=2, value=comp)
            d.alignment = Alignment(horizontal='left')
            d.font = Font(name='Calibri', size=10, bold=False)

            d = dfs.cell(row=row_s, column=3, value=so)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

            d = dfs.cell(row=row_s, column=4, value=su)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

            d = dfs.cell(row=row_s, column=5, value=stot)
            d.alignment = Alignment(horizontal='center')
            d.font = Font(name='Calibri', size=10, bold=False)
            d.number_format = money

        row_s = row_s + 2
        d = dfs.cell(row=row_s, column=2, value='Totals:')
        d.alignment = Alignment(horizontal='right')
        d.font = Font(name='Calibri', size=10, bold=True)

        d = dfs.cell(row=row_s, column=3, value=sumover_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        d = dfs.cell(row=row_s, column=4, value=sumunder_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        d = dfs.cell(row=row_s, column=5, value=sumunder_all+sumover_all)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
        d.number_format = money

        column_widths = column_wide(s_hdrs, zdata)
        for i, column_width in enumerate(column_widths):
            dfs.column_dimensions[get_column_letter(i + 2)].width = column_width + 4

        wb.save(ar_path)

tunnel.stop()