import os
import sys
import socket
from utils import getpaths



import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
from utils import d2s, stripper, hasinput
import datetime
from datetime import timedelta


today = datetime.datetime.today()
today_str = today.strftime("%m/%d/%Y")
d = today.strftime("%B %d, %Y")
cutoff = datetime.datetime.now() - timedelta(30)
cutoff = cutoff.date()
over30 = datetime.datetime.now() - timedelta(30)
over30 = over30.date()
todaydate = today.date()

scac = 'NEVO'
nt = 'remote'
scac = scac.upper()

if scac == 'OSLM' or scac == 'FELA' or scac == 'NEVO':

    print(f'Running FFF_Nevo_Report for {scac}')
    host_name = socket.gethostname()
    print("Host Name:", host_name)
    dropbox_path = getpaths(host_name, 'dropbox')
    ar_path = f'{dropbox_path}/Dray/{scac}_AR_Report.xlsx'
    sys_path = getpaths(host_name, 'system')
    sys.path.append(sys_path) #So we can import CCC_system_setup from full path

    os.environ['SCAC'] = scac
    os.environ['PURPOSE'] = 'script'
    os.environ['MACHINE'] = host_name
    os.environ['TUNNEL'] = nt

    from remote_db_connect import db
    if nt == 'remote': from remote_db_connect import tunnel
    from models8 import Orders, Interchange, Invoices, Openi

else:
    scac = 'nogo'
    print('The argument must be FELA or OSLM or NEVO')
    quit()

if scac != 'nogo':
    wb = openpyxl.load_workbook(ar_path, data_only=True)

    sheet_str = today_str.replace('/',' ')
    newsheet = f'AR Special {sheet_str}'
    dfs= wb.create_sheet(newsheet)

    #formats for writing to excel
    money = '$#,##0.00'
    dec2 = '#,##0.00'
    dec0 = '#,##0'

    #Create column width calculation function
    def column_wide(headers,ydata):
        column_widths = []
        for cell in enumerate(headers):
            cell = str(cell)
            column_widths.append(len(cell))
        for row in ydata:
            for i, cell in enumerate(row):
                test = len(str(cell))
                column_widths[i] = max(test,column_widths[i])
        return column_widths

    #Write the header for Special Report Sheet
    row = 1
    ydata = []
    hdrs = ['SID', 'JO','Company', 'Order', 'Release','Booking In','Container','Date Out','Date In','Date Invoiced', 'Amt Tot','Invoice Filename', 'Driver', 'Truck', 'Amount LH', 'Comment ID']
    for col, hdr in enumerate(hdrs):
        d = dfs.cell(row=row, column=col + 1, value=hdr)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=True)
    success = 0
    trys = 0
    while success == 0 and trys < 20:
        try:
            #odata = Orders.query.filter((Orders.Istat>0) & ((Orders.Istat<4) | (Orders.Istat == 7) | (Orders.Istat == 6)) & (Orders.Date > cutoff)).all()
            odata = Orders.query.filter((Orders.Hstat > 0) & (Orders.Date > cutoff)).all()
            success = 1
        except:
            print(f'Could not open tunnel on try {trys}')
            trys = trys + 1

    if success == 1:
        #Make a list of each job done by Trucker
        comps = []
        for odat in odata:
            comp = odat.Shipper
            if comp not in comps: comps.append(comp)

        for gdat in odata:
            print(gdat.id, gdat.Jo, gdat.Date)
            jo = gdat.Jo
            links = gdat.Links
            odr = gdat.Order
            if links is None: links = ''
            invo = gdat.Invoice
            if invo is None: invo = ''

            obk = gdat.Booking
            obol = gdat.BOL
            if hasinput(obol):
                ietyp = 'imp'
            else:
                ietyp = 'exp'


            int_data = Interchange.query.filter(Interchange.Jo == jo).all()
            if len(int_data) >= 1:
                bk1 = int_data[0].Release
                con = int_data[0].Container
                dt1 = int_data[0].Date
                drv = int_data[0].Driver
                trk = int_data[0].TruckNumber
            else:
                bk1 = 'None'
                dt1 = None
                drv = 'None'
            if len(int_data) >= 2:
                bk2 = int_data[1].Release
                dt2 = int_data[1].Date
            else:
                bk2 = 'None'
                dt2 = None

            if ietyp == 'imp': bk1 = obol

            if bk1 == 'DP OP' or bk2 == 'DP OP':
                print(f'This booking {bk1} and {bk2} has DP OP')
            else:
                idat = Invoices.query.filter(Invoices.Jo == jo).first()
                if idat is not None:
                    dti = idat.Date
                    amt1 = float(idat.Amount)
                    amt2 = float(idat.Total)
                else:
                    dti = 'Not Invoiced'
                    amt1 = float(gdat.Amount)
                    amt2 = 0.00
                newblock = [gdat.id, gdat.Jo, gdat.Shipper, odr, bk1, bk2, con, dt1, dt2, dti, amt2, invo, drv, trk, amt1, links]
                ydata.append(newblock)

        #ydata.sort(key=lambda row: (row[9], row[12]))

        for ydat in ydata:
            row = row + 1
            dti = ydat[9]

            for col, item in enumerate(ydat):
                d = dfs.cell(row=row, column=col + 1, value=item)
                d.alignment = Alignment(horizontal='center')

                if col == 10: d.number_format = money

        column_widths = column_wide(hdrs, ydata)
        for i, column_width in enumerate(column_widths):
            dfs.column_dimensions[get_column_letter(i + 1)].width = column_width + 4


        wb.save(ar_path)

tunnel.stop()
