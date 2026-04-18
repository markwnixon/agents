import os
import sys
import socket
from utils import getpaths, hasinput, d2s, stripper
import paramiko
from pathlib import Path
import time

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
import datetime
from datetime import timedelta
import numpy as np

# Date Section: Define Dates of Interest
# Set up calculation of work days, holidays, etc so that we can make projections
################################################################################
################ Define the Year to Date Stop Point ############################
start_date = datetime.date(2024, 7, 1)
end_date = datetime.date(2024, 12, 31)
################################################################################
categories = ['Line Item Total', 'Line Haul', 'Chassis Fees', 'Detention', 'Storage', 'Demurrage', 'Loading Charge', 'Towing', 'Other']
scac = 'fela'
nt = 'remote'
scac = scac.upper()
print(f'Running Spreadsheet Maker for {scac} from {start_date} to {end_date}')
host_name = socket.gethostname()
print("Host Name:", host_name)
dropbox_path = getpaths(host_name, 'dropbox')
ar_path = f'{dropbox_path}/Dray/{scac}_AR_Report.xlsx'
sys_path = getpaths(host_name, 'system')
sys.path.append(sys_path)  # So we can import CCC_system_setup from full path

os.environ['SCAC'] = scac
os.environ['PURPOSE'] = 'script'
os.environ['MACHINE'] = host_name
os.environ['TUNNEL'] = nt

from remote_db_connect import db
if nt == 'remote':
    from remote_db_connect import tunnel

def getfiles(nt, dir, ssh):
    if nt == 'local':
        file_list = os.listdir(dir)
    if nt == 'remote':
        command = f'ls -a {dir}'
        stdin, stdout, stderr = ssh.exec_command(command)
        file_list = stdout.read().decode().splitlines()
    return file_list

def getdrivers(odat):
    jo = odat.Jo
    driver_in = 'No Driver Found'
    driver_out = 'No Driver Found'
    idata = Interchange.query.filter(Interchange.Jo == jo).all()
    for idat in idata:
        typ = idat.Type
        if 'In' in typ:
            driver_in = idat.Driver
        elif 'Out' in typ:
            driver_out = idat.Driver
    return driver_in, driver_out

from CCC_system_setup import tup, dbp
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(tup[0], 22, tup[1], tup[2])
print(tup[0], tup[1], tup[2])
dirbase = f'{scac.lower()}/webapp/static/{scac}/data'
pkg_dir = f'{dirbase}/vPackage'
pkg_file_list = getfiles(nt, pkg_dir, ssh)

# Crate the output directory
writepath = '/home/mark/Documents/SpreadsheetMaker'
path = Path(writepath)
path.mkdir(exist_ok=True)
print("Directory is ready:", path.resolve())

locrec = f'/home/mark/Documents/SpreadsheetMaker/FELA_Summons_Data.xlsx'
wb = Workbook()
ws = wb.create_sheet(title='Data')
ydata = []

def column_wide(headers, ydata, other):
    column_widths = []
    for i, cell in enumerate(headers):
        cell = str(cell)
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]
    for row in ydata:
        for i, cell in enumerate(row):
            cell = str(cell)
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
    for row in other:
        for i, cell in enumerate(row):
            cell = str(cell)
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
    return column_widths

start_row = 1
# write the headers
headers = ['Delivery Date', 'Delivery Address', 'Shipper', 'Release', 'Container', 'Haul Type', 'Port', 'Gate Out', 'Gate In', 'Driver-Out', 'Driver-In', 'Document Filename']
for jx, col in enumerate(headers):
    d = ws.cell(row=start_row, column=jx + 1, value=col)
    d.font = Font(name='Calibri', size=10, bold=True)
    d.alignment = Alignment(horizontal='center')




from models8 import Orders, Drops, Gledger, Invoices, Interchange, Income, People

success = 0
trys = 0
while success == 0:
    try:
        odata = Orders.query.filter((Orders.Date3 >= start_date) & (Orders.Date3 <= end_date)).order_by(Orders.Date3).all()
        success = 1
    except:
        print(f'No success opening tunnel on try {trys}')
        success = 0
        trys = trys + 1
        if trys > 10:
            if tunnel: tunnel.stop()
            quit()



for odat in odata:
    ht = odat.HaulType
    dp1 = odat.Dropblock1
    sh = odat.Shipper
    driver_in, driver_out = getdrivers(odat)
    if dp1 is None: port = 'No Dropblock1'
    else:
        dp1 = dp1.lower()
        if'baltimore' in dp1 or 'seagirt' in dp1:
            port = 'Baltimore'
        elif 'nit' in dp1 or 'vig' or 'virginia' in dp1 or 'portsmouth' in dp1:
            port = 'Virginia'
        elif 'packer' in dp1 or 'philadelphia' in dp1:
            port = 'Philadelphia'
        else:
            port = 'None'
    #print(f'Date: {odat.Date3}, Container: {odat.Container}, HaulType: {odat.HaulType} Port: {port}')

    if 'Dray' in ht and port == 'Baltimore':
        print(f'Delivery Date: {odat.Date3}, Delivery Address: {odat.Dropblock2} Shipper: {sh}. Release: {odat.Booking} Container: {odat.Container}, HaulType: {odat.HaulType} Port: {port} Gate Out: {odat.Date} Gate In: {odat.Date2} Driver-Outgate: {driver_out}, Driver-Ingate: {driver_in}')

        # Check if there is a package file
        pfile = odat.Package
        if pfile is None:
            if sh == 'FEL Ocean Div':
                print('FEL Ocean Job so No package exist for this job')
            else:
                print('******************************')
                print('Why is there no Package for this job?')
                print('******************************')
        elif pfile in pkg_file_list:
            print(f'Package file {pfile} found')
            fpath = f'{pkg_dir}/{pfile}'
            newwritepath = f'{writepath}/{pfile}'
            if os.path.exists(newwritepath):
                print(f'Package file {pfile} already exists on local machine')
            else:
                print(f'Attempting to copy remote file {fpath} to local directory {newwritepath}')
                max_tries = 5
                for attempt in range(max_tries):
                    try:
                        transport = paramiko.Transport((tup[0], 22))
                        transport.connect(username=tup[1], password=tup[2])
                        print('Connected')
                        break
                    except paramiko.SSHException as e:
                        print(f"Attempt {attempt + 1} failed: {e}")
                        time.sleep(5)  # Wait before retrying
                    else:
                        raise Exception("Failed to connect after several attempts")

                sftp = paramiko.SFTPClient.from_transport(transport)
                sftp.get(fpath, newwritepath)
                sftp.close()
                transport.close()


        else:
            print(f'Package file {pfile} NOT found')

        ydata.append(
            [odat.Date3, odat.Dropblock2, sh, odat.Booking, odat.Container, odat.HaulType, port, odat.Date, odat.Date2,
             driver_out, driver_in, pfile])


print('The full package file list:')
print(pkg_file_list)

money = '$#,##0.00'
dec2 = '#,##0.00'
dec0 = '#,##0'
letcol = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
row = start_row
flist = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
for mdata in ydata:
    row = row + 1
    offset = 0
    for ix, col in enumerate(mdata):
        jx = ix + offset + 1
        d = ws.cell(row=row, column=jx, value=col)
        d.alignment = Alignment(horizontal='center')
        d.font = Font(name='Calibri', size=10, bold=False)
        if flist[ix] != 0: d.number_format = flist[ix]

other = []
column_widths = column_wide(headers, ydata, other)
for i, column_width in enumerate(column_widths): ws.column_dimensions[get_column_letter(i+1)].width = column_width + 4
wb.save(locrec)








tunnel.stop()

